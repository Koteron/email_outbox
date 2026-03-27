import logging
from typing import Iterator, Dict
from openpyxl import load_workbook
from django.db import IntegrityError, transaction
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from outbox.models import OutboxRecord


logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    "external_id",
    "user_id",
    "email",
    "subject",
    "message",
]


def _iter_xlsx_rows(file_path: str) -> Iterator[Dict]:
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)

    headers = next(rows)
    header_map = {h.strip(): i for i, h in enumerate(headers)}

    for col in REQUIRED_COLUMNS:
        if col not in header_map:
            raise ValueError(f"Missing required column: {col}")

    for row in rows:
        yield {col: row[header_map[col]] for col in REQUIRED_COLUMNS}


def import_outbox_records(file_path: str, batch_size: int = 500):

    total = 0
    created = 0
    skipped = 0
    failed = 0

    batch = []

    for row in _iter_xlsx_rows(file_path):
        total += 1

        try:
            validate_email(str(row["email"]))
            record = OutboxRecord(
                external_id=row["external_id"],
                user_id=row["user_id"],
                email=row["email"],
                subject=row["subject"],
                message=row["message"],
            )
            batch.append(record)

            if len(batch) >= batch_size:
                logger.debug("Batch size reached, flushing...")
                c, s = _flush_batch(batch)
                created += c
                skipped += s
                batch.clear()
                logger.debug(f"Flushed {c} and skipped {s}")
        
        except ValidationError:
            logger.debug("Invalid email skipped", extra={"email": row["email"]})
            failed += 1
        
        except Exception:
            logger.exception("Row processing failed", extra={"row": row})
            failed += 1

    if batch:
        c, s = _flush_batch(batch)
        created += c
        skipped += s

    return {
        "total": total,
        "created": created,
        "skipped": skipped,
        "failed": failed,
    }


def _flush_batch(batch):
    created = 0
    skipped = 0

    try:
        with transaction.atomic():
            objs = OutboxRecord.objects.bulk_create(
                batch,
            )
            created = len(objs)
            skipped = len(batch) - created

    except Exception:
        created = 0
        skipped = 0

        for obj in batch:
            try:
                obj.save()
                created += 1
            except IntegrityError:
                skipped += 1
            except Exception:
                logger.exception("Failed to save record", extra={"obj": obj})
                skipped += 1

    return created, skipped